# python imports
import datetime

# django imports
from django.utils.translation import gettext_lazy as _
from django.shortcuts import HttpResponse, redirect

# 3rd party imports
from oscar.core.loading import get_class, get_model
from openpyxl import Workbook

# internal imports.
ReportGenerator = get_class('dashboard.reports.reports', 'ReportGenerator')
ReportCSVFormatter = get_class('dashboard.reports.reports', 'ReportCSVFormatter')
ReportHTMLFormatter = get_class('dashboard.reports.reports', 'ReportHTMLFormatter')

Order = get_model('order', 'Order')

Partner = get_model('partner', 'Partner')
StockRecord = get_model('partner', 'StockRecord')

Basket_Line = get_model('basket', 'Line')
Basket = get_model('basket', 'Basket')


ProductRecord = get_model('analytics', 'ProductRecord')
UserRecord = get_model('analytics', 'UserRecord')

Product = get_model('catalogue', 'Product')


class OrderReportCSVFormatter(ReportCSVFormatter):

    """
    Oscar extended method to download reports.
    """

    filename_template = 'orders-%s-to-%s.csv'

    def generate_response(self, objects, **kwargs):

        """
        Method to download report as xls.
        :param objects:
        :param kwargs:
        :return:
        """

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return self.generate_xls(response, objects)

    def generate_xls(self, response, orders):

        """
        Method to generate xls.
        :param response:
        :param orders:
        :return:
        """

        filename = 'order_%s' % (str(datetime.datetime.now()))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        excel_data = [
            [
                'Sr. No', 'Order Number', 'Order value', 'Date placed',
                'Address', 'Order status'
            ]
        ]

        counter = 0

        for data in orders:
            counter = counter + 1

            excel_data.append(
                [
                    str(counter), str(data.number), str(data.total_incl_tax),
                    str(data.date_placed.date()), str(data.shipping_address),
                    str(data.status)
                ]
            )

        for line in excel_data:
            ws.append(line)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % (filename)

        wb.save(response)

        return response

    def generate_csv(self, response, orders):

        writer = self.get_csv_writer(response)
        header_row = [
            _('Order number'), _('Name'),
            _('Email'), _('Total incl. tax'),
            _('Date placed')
        ]
        writer.writerow(header_row)

        for order in orders:
            row = [
                order.number,
                '-' if order.is_anonymous else order.user.get_full_name(),
                order.email,
                order.total_incl_tax,
                self.format_datetime(order.date_placed)
            ]
            writer.writerow(row)

    def filename(self, **kwargs):
        return self.filename_template % (kwargs['start_date'], kwargs['end_date'])


class OrderReportHTMLFormatter(ReportHTMLFormatter):

    """
    Oscar extended report html formatter.
    """

    filename_template = 'dashboard/reports/partials/order_report.html'


class OrderReportGenerator(ReportGenerator):

    """
    Oscar extended order report.
    """

    code = 'order_report'
    description = _("Orders placed")
    date_range_field_name = 'date_placed'

    formatters = {
        'CSV_formatter': OrderReportCSVFormatter,
        'HTML_formatter': OrderReportHTMLFormatter,
    }

    def generate(self,user):
        isVendor = user.groups.filter(name='Vendor').exists()
        if isVendor:
            partner_id = Partner.objects.get(users=user)
            stock_record = StockRecord.objects.filter(partner_id=partner_id.id).values_list('product_id')
            basket_line = Basket_Line.objects.filter(product_id__in=stock_record).values_list('basket_id')
            bakset_obj = Basket.objects.filter(id__in=basket_line).values_list('id')
            orders = Order.objects.filter(basket_id__in=bakset_obj)
            qs = orders
        else:
            qs = Order._default_manager.all()
        if self.start_date:
            qs = qs.filter(date_placed__gte=self.start_date)
        if self.end_date:
            qs = qs.filter(
                date_placed__lt=self.end_date + datetime.timedelta(days=1))

        additional_data = {
            'start_date': self.start_date,
            'end_date': self.end_date
        }

        return self.formatter.generate_response(qs, **additional_data)

    def is_available_to(self, user):
        return user.is_staff


class ProductReportCSVFormatter(ReportCSVFormatter):

    """
    Oscar extended product reports.
    """

    filename_template = 'conditional-offer-performance.csv'

    def generate_response(self, objects, **kwargs):

        """
        Method to download report as xls.
        :param objects:
        :param kwargs:
        :return:
        """

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return self.generate_xls(response, objects)

    def generate_xls(self, response, products):

        """
        Method to generate xls.
        :param response:
        :param orders:
        :return:
        """

        filename = 'product_%s' % (str(datetime.datetime.now()))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        excel_data = [
            [
                'Sr. No', 'Product', 'Views', 'Basket additions',
                'Purchases'
            ]
        ]

        counter = 0

        for data in products:
            counter = counter + 1

            excel_data.append(
                [
                    str(counter), str(data.product), str(data.num_views),
                    str(data.num_basket_additions), str(data.num_purchases)
                ]
            )

        for line in excel_data:
            ws.append(line)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % (filename)

        wb.save(response)

        return response

    def generate_csv(self, response, products):

        writer = self.get_csv_writer(response)
        header_row = [
            _('Product'), _('Views'),
            _('Basket additions'), _('Purchases')
        ]
        writer.writerow(header_row)

        for record in products:
            row = [
                record.product, record.num_views,
                record.num_basket_additions, record.num_purchases
            ]
            writer.writerow(row)


class ProductReportHTMLFormatter(ReportHTMLFormatter):

    """
    Oscar extend method to for product html.
    """

    filename_template = 'dashboard/reports/partials/product_report.html'


class ProductReportGenerator(ReportGenerator):

    """
    Oscar extend method to format product report.
    """

    code = 'product_analytics'
    description = _('Product analytics')

    formatters = {
        'CSV_formatter': ProductReportCSVFormatter,
        'HTML_formatter': ProductReportHTMLFormatter}

    def report_description(self):
        return self.description

    def generate(self, user):
        isVendor = user.groups.filter(name='Vendor').exists()
        if isVendor:
            partner_obj = Partner.objects.get(users=user)
            stock_obj = StockRecord.objects.filter(partner=partner_obj).values_list('product_id')
            product_objs = Product.objects.filter(id__in=stock_obj,is_deleted=False).values_list('id')
            product_analytics = ProductRecord.objects.filter(product_id__in=product_objs)
            records =  product_analytics
        else:
            records = ProductRecord.objects.all()
        return self.formatter.generate_response(records)

    def is_available_to(self, user):
        return user.is_staff


class UserReportCSVFormatter(ReportCSVFormatter):

    """
    Oscar extended method to format user report.
    """

    filename_template = 'user-analytics.csv'

    def generate_csv(self, response, users):

        writer = self.get_csv_writer(response)
        header_row = [
            _('Name'), _('Date registered'),
            _('Product views'), _('Basket additions'),
            _('Orders'), _('Order lines'),
            _('Order items'), _('Total spent'),
            _('Date of last order')
        ]
        writer.writerow(header_row)

        for record in users:
            row = [
                record.user.get_full_name(), self.format_date(record.user.date_joined),
                record.num_product_views, record.num_basket_additions,
                record.num_orders, record.num_order_lines,
                record.num_order_items, record.total_spent,
                self.format_datetime(record.date_last_order)
            ]
            writer.writerow(row)


class UserReportHTMLFormatter(ReportHTMLFormatter):

    """
    Oscar extended method to format user report html.
    """

    filename_template = 'dashboard/reports/partials/user_report.html'


class UserReportGenerator(ReportGenerator):

    """
    Oscar extended method to for user reports.
    """

    code = 'user_analytics'
    description = _('User analytics')

    formatters = {
        'CSV_formatter': UserReportCSVFormatter,
        'HTML_formatter': UserReportHTMLFormatter
    }

    def generate(self):
        users = UserRecord._default_manager.select_related().all()
        return self.formatter.generate_response(users)

    def is_available_to(self, user):
        return user.is_staff


class OpenBasketReportCSVFormatter(ReportCSVFormatter):

    """
    Oscar open basket report.
    """

    filename_template = 'open-baskets-%s-%s.csv'

    def generate_csv(self, response, baskets):

        writer = self.get_csv_writer(response)
        header_row = [
            _('User ID'), _('Name'),
            _('Email'), _('Basket status'),
            _('Num lines'), _('Num items'),
            _('Date of creation'), _('Time since creation'),
        ]
        writer.writerow(header_row)

        for basket in baskets:
            if basket.owner:
                row = [basket.owner_id, basket.owner.get_full_name(),
                       basket.owner.email,
                       basket.status, basket.num_lines,
                       self.format_datetime(basket.date_created),
                       basket.time_since_creation]
            else:
                row = [basket.owner_id, None, None, basket.status,
                       basket.num_lines, basket.num_items,
                       self.format_datetime(basket.date_created),
                       basket.time_since_creation]
            writer.writerow(row)

    def filename(self, **kwargs):
        return self.filename_template % (
            kwargs['start_date'],kwargs['end_date']
        )


class OpenBasketReportHTMLFormatter(ReportHTMLFormatter):

    """
    Oscar method to open basket html formatter.
    """

    filename_template = 'dashboard/reports/partials/open_basket_report.html'


class OpenBasketReportGenerator(ReportGenerator):

    """
    Report of baskets which haven't been submitted yet
    """

    code = 'open_baskets'
    description = _('Open baskets')
    date_range_field_name = 'date_created'

    formatters = {
        'CSV_formatter': OpenBasketReportCSVFormatter,
        'HTML_formatter': OpenBasketReportHTMLFormatter}

    def generate(self):
        additional_data = {
            'start_date': self.start_date,
            'end_date': self.end_date
        }
        baskets = Basket._default_manager.filter(status=Basket.OPEN)
        return self.formatter.generate_response(baskets, **additional_data)


class SubmittedBasketReportCSVFormatter(ReportCSVFormatter):

    """
    Oscar extended method to generate submitted basket.
    """

    filename_template = 'submitted_baskets-%s-%s.csv'

    def generate_response(self, objects, **kwargs):

        """
        Method to download report as xls.
        :param objects:
        :param kwargs:
        :return:
        """

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return self.generate_xls(response, objects)

    def generate_xls(self, response, baskets):

        """
        Method to generate xls.
        :param response:
        :param orders:
        :return:
        """

        filename = 'baskets_submitted_%s' % (str(datetime.datetime.now()))

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        excel_data = [
            [
                'Sr. No', 'Email', 'Num of items', 'Date created'
            ]
        ]

        counter = 0

        for data in baskets:
            counter = counter + 1

            excel_data.append(
                [
                    str(counter), str(data.owner), str(data.num_items),
                    self.format_datetime(data.date_created),
                ]
            )

        for line in excel_data:
            ws.append(line)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % (filename)

        wb.save(response)

        return response

    def generate_csv(self, response, baskets):

        writer = self.get_csv_writer(response)
        header_row = [
            _('User ID'), _('User'), _('Basket status'), _('Num lines'),
            _('Num items'), _('Date created'), _('Time between creation and submission'),
        ]
        writer.writerow(header_row)

        for basket in baskets:
            row = [
                basket.owner_id,
                basket.owner,
                basket.status,
                basket.num_lines,
                basket.num_items,
                self.format_datetime(basket.date_created),
                basket.time_before_submit
            ]
            writer.writerow(row)

    def filename(self, **kwargs):
        return self.filename_template % (kwargs['start_date'], kwargs['end_date'])


class SubmittedBasketReportHTMLFormatter(ReportHTMLFormatter):

    """
    Oscar method to sub. basket html formatter.
    """

    filename_template = 'dashboard/reports/partials/submitted_basket_report.html'


class SubmittedBasketReportGenerator(ReportGenerator):

    """
    Report of baskets that have been submitted
    """

    code = 'submitted_baskets'
    description = _('Submitted baskets')
    date_range_field_name = 'date_submitted'

    formatters = {
        'CSV_formatter': SubmittedBasketReportCSVFormatter,
        'HTML_formatter': SubmittedBasketReportHTMLFormatter}

    def generate(self,user):
        additional_data = {
            'start_date': self.start_date,
            'end_date': self.end_date
        }

        isVendor = user.groups.filter(name='Vendor').exists()
        if isVendor:
            partner_id = Partner.objects.get(users=user)
            stock_record = StockRecord.objects.filter(partner_id=partner_id.id).values_list('product_id')
            basket_line = Basket_Line.objects.filter(product_id__in=stock_record).values_list('basket_id')
            bakset_obj = Basket.objects.filter(id__in=basket_line)
            baskets = bakset_obj
        else:
            baskets = Basket._default_manager.filter(status=Basket.SUBMITTED)
        return self.formatter.generate_response(baskets, **additional_data)
